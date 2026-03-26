import streamlit as st
import pandas as pd
import openpyxl
import io
import json
from google import genai

# --- 1. 페이지 설정 및 보안 설정 ---
st.set_page_config(page_title="금융 데이터 감사 통합 툴킷", layout="wide")

# --- 2. AI 설정 (Gemini API) ---
GENAI_API_KEY = "AIzaSyBNjveTdD_-K20-713IyOYJtTrjKpU95bs"
client = genai.Client(api_key=GENAI_API_KEY)
MODEL_ID = 'gemini-2.0-flash'

# --- 3. 사이드바 메뉴 구성 ---
st.sidebar.title("🛠️ Audit Menu")
menu = st.sidebar.selectbox(
    "기능을 선택하세요",
    ["메인 화면", "1) 수식 추출", "2) 오타 검정 (AI 검사)", "3) 병합 해제 (데이터 클렌징)"]
)

st.sidebar.markdown("---")
st.sidebar.info("💡 **Tip**: 2번 AI 검사 시 '행 수'를 조절하여 속도를 관리하세요.")

# --- 기능별 상세 구현 ---

# [메인 화면]
if menu == "메인 화면":
    st.title("🏦 금융 데이터 감사 및 클렌징 통합 시스템")
    st.write("""
    이 시스템은 금융권 실무자의 업무 효율성을 높이기 위해 세 가지 핵심 기능을 제공합니다.
    
    1. **수식 추출**: 모든 시트를 스캔하여 셀에 걸린 수식(=)을 찾아 연결 시트를 검토합니다.
    2. **오타 검정**: 강력한 AI 프롬프트를 활용해 실무 데이터 속 오타를 정밀 검수합니다.
    3. **병합 해제**: 수식 참조 오류 방지를 위해 병합된 셀을 풀고 데이터를 자동 채웁니다.
    """)

# [기능 1: 수식 추출]
elif menu == "1) 수식 추출":
    st.header("🔍 수식 위치 및 원문 리스트업")
    st.write("부서장 운영효과성 검토를 위해 파일 내 모든 수식을 스캔합니다.")
    
    file = st.file_uploader("엑셀 파일(.xlsx) 업로드", type=["xlsx"], key="formula_upload")
    
    if file:
        if st.button("수식 스캔 실행"):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=False)
                results = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows():
                        for cell in row:
                            val = cell.value
                            if cell.data_type == 'f' or (isinstance(val, str) and str(val).startswith('=')):
                                results.append({
                                    "시트명": sheet,
                                    "셀주소": cell.coordinate,
                                    "수식내용": str(val)
                                })
                
                if results:
                    df_res = pd.DataFrame(results)
                    st.success(f"✅ 총 {len(results)}개의 수식을 발견했습니다.")
                    st.dataframe(df_res, use_container_width=True)
                    
                    output = io.BytesIO()
                    df_res.to_excel(output, index=False)
                    st.download_button("수식 리포트 다운로드", output.getvalue(), "formula_report.xlsx")
                else:
                    st.info("🔍 수식이 포함된 셀이 없습니다.")
            except Exception as e:
                st.error(f"❌ 오류 발생: {e}")

# [기능 2: 오타 검정 - 정밀 검수 및 다중 컬럼 지원]
elif menu == "2) 오타 검정 (AI 검사)":
    st.header("🤖 AI 기반 텍스트 오타 정밀 검정")
    st.write("선택한 컬럼에서 오타, 깨진 글자, 띄어쓰기 오류를 집중적으로 찾아냅니다.")
    
    # 업로드 창이 사라지지 않도록 위치 확인
    file = st.file_uploader("검수할 엑셀 파일을 업로드하세요", type=["xlsx"], key="typo_upload")
    
    if file:
        df = pd.read_excel(file)
        target_cols = st.multiselect("검사할 컬럼(열)을 선택하세요", options=df.columns, default=[df.columns[0]])
        
        # 데이터 행 수 슬라이더 (최소 20행으로 설정)
        max_rows = len(df)
        num_rows = st.slider("검사할 행 수", 1, max_rows, min(20, max_rows))
        
        if st.button("AI 검수 시작"):
            if not target_cols:
                st.error("⚠️ 검사할 컬럼을 하나 이상 선택해 주세요.")
            else:
                errors_all = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for idx in range(num_rows):
                    status_text.text(f"⏳ 전체 {num_rows}행 중 {idx+1}번째 행 분석 중...")
                    
                    # 선택된 컬럼들만 데이터 추출하여 JSON 변환
                    row = df.iloc[idx]
                    row_data = {col: row[col] for col in target_cols}
                    row_json = json.dumps(row_data, ensure_ascii=False)
                    
                    # 사용자 검증 완료된 강력한 프롬프트
                    prompt = f"""
                    당신은 텍스트에서 '깨진 글자'와 '명백한 맞춤법 오류'만 찾아내는 기계입니다.
                    
                    [데이터]
                    {row_json}

                    [검수 규칙 - 이 조건이 아니면 절대로 보고하지 마세요]
                    1. 자모음 분리: 글자 끝에 자음/모음이 남은 경우
                    2. 명백한 오타: 받침이 틀렸거나 자음이 잘못 들어간 경우
                    3. 띄어쓰기 오류: 단어 중간에 공백이 들어간 경우

                    [절대 금지 - 위반 시 업무 실패]
                    - 정상적인 한글 단어를 오타라고 보고하지 마세요.
                    - 원문(Original)과 추천(Suggestion)이 똑같은 경우는 절대로 보고하지 마십시오.
                    - 문맥을 '개선'하려고 하지 마십시오. 틀린 '글자'만 찾으십시오.

                    [응답 형식]
                    - 오류가 있을 때만 JSON 출력. 없으면 반드시 {{"errors": []}}
                    
                    {{
                        "errors": [
                            {{
                                "column": "컬럼명", 
                                "original": "틀린 글자", 
                                "suggestion": "바른 글자", 
                                "reason": "오타/자모음분리/띄어쓰기 중 택1"
                            }}
                        ]
                    }}
                    """

                    try:
                        response = client.models.generate_content(
                            model=MODEL_ID,
                            contents=prompt,
                            config={'response_mime_type': 'application/json', 'temperature': 0.0}
                        )
                        errs = json.loads(response.text).get("errors", [])
                        for e in errs:
                            orig = str(e.get('original', '')).strip()
                            sugg = str(e.get('suggestion', '')).strip()
                            # 원문과 추천이 다르고 비어있지 않은 경우만 수집
                            if orig and sugg and orig != sugg:
                                e['row_index'] = idx + 2 # 엑셀 실제 행 번호
                                errors_all.append(e)
                    except:
                        continue
                    
                    progress_bar.progress((idx + 1) / num_rows)
                
                status_text.empty()
                
                if errors_all:
                    report_df = pd.DataFrame(errors_all)
                    st.success(f"✅ 검수 완료! {len(errors_all)}개의 확실한 오타를 발견했습니다.")
                    
                    # 테이블 출력
                    display_cols = ['row_index', 'column', 'original', 'suggestion', 'reason']
                    st.dataframe(report_df[display_cols], use_container_width=True)
                    
                    # 결과 파일 다운로드
                    output = io.BytesIO()
                    report_df.to_excel(output, index=False)
                    st.download_button("오타 리포트 다운로드", output.getvalue(), "typo_report_fixed.xlsx")
                else:
                    st.success("🎉 발견된 오타가 없습니다. 데이터가 아주 깨끗합니다.")

# [기능 3: 병합 해제]
elif menu == "3) 병합 해제 (데이터 클렌징)":
    st.header("🔓 셀 병합 해제 및 데이터 자동 채우기")
    st.write("병합을 해제하고 모든 빈칸에 값을 채워 수식 참조 오류를 방지합니다.")
    
    file = st.file_uploader("원본 엑셀 파일을 업로드하세요", type=["xlsx"], key="unmerge_upload")
    
    if file:
        if st.button("데이터 클렌징 실행"):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                count = 0
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    merged_ranges = list(ws.merged_cells.ranges)
                    for r in merged_ranges:
                        min_col, min_row, max_col, max_row = r.bounds
                        val = ws.cell(row=min_row, column=min_col).value
                        ws.unmerge_cells(str(r))
                        for row_idx in range(min_row, max_row + 1):
                            for col_idx in range(min_col, max_col + 1):
                                ws.cell(row=row_idx, column=col_idx).value = val
                        count += 1
                
                out = io.BytesIO()
                wb.save(out)
                st.success(f"✅ 성공! {count}개의 병합 영역을 해제하고 데이터를 채웠습니다.")
                st.download_button("클렌징된 파일 다운로드", out.getvalue(), "cleansed_data.xlsx")
            except Exception as e:
                st.error(f"❌ 오류 발생: {e}")